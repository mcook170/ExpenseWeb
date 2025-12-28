from flask import Flask, render_template, request, redirect, flash, session, url_for, send_file 
from datetime import datetime, timedelta 
import openpyxl 
import os 
from dotenv import load_dotenv 
from pathlib import Path 
from models import db, User, Expense  # ← import User model
import shutil
import pandas as pd

from werkzeug.security import generate_password_hash, check_password_hash

# Load .env
env_path = Path(__file__).parent / "expensive_stuff.env" 
load_dotenv(dotenv_path=env_path)
app = Flask(__name__) 
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///tracker.db' 
app.config['SECRET_KEY'] = os.environ.get("TOKEN", "dev-secret") 
app.permanent_session_lifetime = timedelta(minutes=15)

db.init_app(app)

with app.app_context(): 
    db.create_all()

# Defines to_dict() function for showing recent expenses
def to_dict(self):
    return {
        "date": str(self.date),  # if date is a datetime object
        "category": self.category,
        "amount": float(self.amount)  # make sure this is a float
    }

# 🧾 Registration Route
@app.route("/register", methods=["GET", "POST"]) 
def register(): 
    if request.method == "POST": 
        username = request.form["username"] 
        pin = request.form["pin"]
        
        if User.query.filter_by(username=username).first():
            flash("Username already exists.")
            return redirect(url_for("register"))

        user = User(username=username, pin_hash=generate_password_hash(pin))
        db.session.add(user)
        db.session.commit()
        flash("Account created. Please log in.")

        template_path = os.path.join(app.root_path, "expenses-template.xlsx")
        user_file = os.path.join(app.root_path, f"{username}_expenses.xlsx")

        if not os.path.exists(template_path):
            flash("Template file is missing. Please contact support.")
            return redirect(url_for("register"))
        if not os.path.exists(user_file):
            shutil.copy(template_path, user_file)


        return redirect(url_for("login"))

    return render_template("register.html", theme=session.get("theme","desert-theme"))

def get_user_filepath(username):
    if "username" not in session:
        flash("Session expired. Please log in.")
        return redirect(url_for("login"))
    return os.path.join(app.root_path, f"{username}_expenses.xlsx")


# 🔐 Login Route
@app.route("/login", methods=["GET", "POST"]) 
def login(): 
    if request.method == "POST": 
        username = request.form["username"] 
        pin = request.form["pin"] 
        user = User.query.filter_by(username=username).first()

        if user and check_password_hash(user.pin_hash, pin):
            session["user_id"] = user.id
            session["username"] = user.username
            session.permanent = True
            return redirect(url_for("index"))

        flash("Invalid username or PIN.")
        return redirect(url_for("login"))

    return render_template("login.html", theme=session.get("theme","desert-theme"))

# 🧾 Tracker Route
@app.route("/", methods=["GET", "POST"]) 
def index(): 
    if "user_id" not in session: 
        return redirect(url_for("login"))
        
    username = session["username"]
    filename = f"{username}_expenses.xlsx"
    filepath = get_user_filepath(username)

    # Handle form submission
    if request.method == "POST":
        description = request.form.get("description")
        expense_type = request.form.get("expense_type")
        amount = request.form.get("amount")
        note = request.form.get("note")
        date = datetime.now().strftime("%m/%d/%Y")

        try:
            amount = float(amount)
        except (ValueError, TypeError):
            amount = 0.00

        new_expense = Expense(
            user_id=session["user_id"],
            expense_type=expense_type,
            description=description,
            amount=amount,
            note=note
        )

        db.session.add(new_expense)
        db.session.commit()

        return redirect(url_for("index"))

    if request.method=="GET":
        expenses = (
            Expense.query
            .filter_by(user_id=session["user_id"])
            .order_by(Expense.date.desc())
            .limit(10)
            .all()
        )
            
        return render_template("index.html", username=username, entries=expenses, theme=session.get("theme","desert-theme"))

# --- route to receive theme change / uploads ---
@app.route('/set_wallpaper', methods=['POST'])
def set_wallpaper():
    user = get_current_user()
    if not user:
        flash("You must be logged in to change wallpaper", "error")
        return redirect(url_for('index'))

    action_type = request.form.get('type')
    # 1) built-in theme selection
    if action_type == 'theme':
        theme_name = request.form.get('theme_name')
        # store theme class separately (so fonts/colors remain controlled)
        # optionally: user.theme_name column; if not present, set user.theme = theme_name
        user.theme_name = theme_name  # create this column if you don't have it
        # if user had a custom wallpaper file, leave it (or clear it depending on desired behavior)
        db.session.commit()
        flash("Theme changed", "success")
        return redirect(url_for('index'))
    
    # 2) reset to defaults
    if action_type == 'reset':
        # clear custom wallpaper and reset theme to desert
        user.theme_name = 'desert-theme'
        # if you stored path in user.wallpaper_path, clear it
        if hasattr(user, 'wallpaper_path'):
            user.wallpaper_path = None
        db.session.commit()
        flash("Reset to default wallpaper", "success")
        return redirect(url_for('index'))

    # 3) upload a new custom wallpaper
    if action_type == 'upload':
        if 'wallpaper_file' not in request.files:
            flash('No file part', 'error')
            return redirect(url_for('index'))
        file = request.files['wallpaper_file']
        if file.filename == '':
            flash('No selected file', 'error')
            return redirect(url_for('index'))
        if file and allowed_file(file.filename):
            # optional: enforce size (simple check)
            file.seek(0, os.SEEK_END)
            filesize = file.tell()
            file.seek(0)
            if filesize > MAX_UPLOAD_SIZE:
                flash("File too large (max 4MB)", "error")
                return redirect(url_for('index'))

            filename = secure_filename(file.filename)
            # create per-user upload folder inside static/uploads/<username>/
            upload_folder = os.path.join(current_app.static_folder, 'uploads', user.username)
            Path(upload_folder).mkdir(parents=True, exist_ok=True)
            save_path = os.path.join(upload_folder, filename)
            file.save(save_path)

            # store relative path to the static folder (so url_for('static', filename=...) works)
            rel_path = os.path.join('uploads', user.username, filename)
            user.wallpaper_path = rel_path  # create this column if you don't have it
            db.session.commit()

            flash("Wallpaper uploaded", "success")
            return redirect(url_for('index'))
        else:
            flash("Invalid file type. Allowed: png, jpg, jpeg", "error")
            return redirect(url_for('index'))

    # fallback
    flash("Unknown action", "error")
    return redirect(url_for('index'))

# Sheet View (All Expenses)
@app.route("/all_expenses")
def all_expenses():
    if "user_id" not in session:
        flash("Session expired. Please log in.")
        return redirect(url_for("login"))
    
    page = request.args.get("page", 1, type=int)
    expenses = (
        Expense.query
        .filter_by(user_id=session["user_id"])
        .order_by(Expense.date.desc())
        .paginate(page=page, per_page=20)
    )


    return render_template("all_expenses.html", username=session["username"], entries=expenses)


#📊 Download Route
@app.route("/download")
def download():
    if "user_id" not in session:
        flash("Session expired. Please log in.")
        return redirect(url_for("login"))

    username = session["username"]

    # 1. Query from DB
    expenses = (
        Expense.query
        .filter_by(user_id=session["user_id"])
        .order_by(Expense.date.asc())
        .all()
    )

    # 2. Load template (with formulas in Summary sheet)
    template_path = os.path.join(app.root_path, "expenses-template.xlsx")
    wb = openpyxl.load_workbook(template_path)
    sheet = wb["Expenses"]

    # 3. Fill the Expenses sheet (starting at row 5)
    start_row = 5
    for i, expense in enumerate(expenses, start=start_row):
        sheet.cell(row=i, column=1, value=expense.date.strftime("%m/%d/%Y"))
        sheet.cell(row=i, column=2, value=expense.expense_type)
        sheet.cell(row=i, column=3, value=expense.description)
        sheet.cell(row=i, column=4, value=expense.amount)
        sheet.cell(row=i, column=5, value=expense.note)

    # 4. Save to temp file
    filepath = os.path.join(app.root_path, f"{username}_expenses.xlsx")
    wb.save(filepath)

    # 5. Send file back to user
    return send_file(filepath, as_attachment=True, download_name=f"{username}_expenses.xlsx")

# 🗑️ Delete Account Route
@app.route("/delete_account", methods=["POST"])
def delete_account():
    if "user_id" not in session:
        return redirect(url_for("login"))

    user = User.query.get(session["user_id"])
    if user:
        db.session.delete(user)
        db.session.commit()

    user_file = get_user_filepath(session["username"])
    if os.path.exists(user_file):
        os.remove(user_file)

    session.clear()
    flash("Account deleted. You may now re-register.")
    return redirect(url_for("register"))

# Refresh Template Route
@app.route("/refresh_workbook", methods=["POST"])
def refresh_workbook():
    if "username" not in session:
        return redirect(url_for("login"))

    filepath = get_user_filepath(session["username"])
    template_path = os.path.join(app.root_path, "expenses-template.xlsx")

    if os.path.exists(template_path):
        shutil.copy(template_path, filepath)
        flash("Workbook refreshed with the latest template.")
    else:
        flash("Template file is missing. Please contact support.")

    return redirect(url_for("index"))

# 🎨 Theme Route
@app.route("/set_theme/<theme>")
def set_theme(theme):
    # Save the selected theme in session
    session["theme"] = theme

    # Redirect back to the previous page if possible, otherwise home
    return redirect(request.referrer or url_for("index"))

#🚪 Logout
@app.route("/logout") 
def logout(): 
    session.clear() 
    return redirect(url_for("login"))

if __name__ == "__main__":
    app.run(debug=True)